import os
import PyPDF2
from docx import Document
from openpyxl import load_workbook

import google.generativeai as genai




os.environ["GOOGLE_API_KEY"] = "AIzaSyBzjeDIadHpU7--EbmefNQIfoX-w8BWmQM"
# Prompt generation
def generate_text(prompt):
    model = genai.GenerativeModel("gemini-1.5-flash")
    genai.configure(api_key="AIzaSyBzjeDIadHpU7--EbmefNQIfoX-w8BWmQM")
    response = model.generate_content(prompt)
    print(response.text)
    return response.text




def extract_text_from_file(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext == '.pdf':
        try:
            reader = PyPDF2.PdfReader(uploaded_file)
            return '\n'.join([page.extract_text() or '' for page in reader.pages])
        except:
            return ''
    elif ext == '.docx':
        try:
            doc = Document(uploaded_file)
            return '\n'.join([para.text for para in doc.paragraphs])
        except:
            return ''
    elif ext in ['.xlsx', '.xls']:
        try:
            wb = load_workbook(uploaded_file, data_only=True)
            text = ''
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += ' '.join([str(cell) for cell in row if cell]) + '\n'
            return text
        except:
            return ''
    elif ext == '.txt':
        try:
            return uploaded_file.read().decode('utf-8')
        except:
            return ''
    else:
        return ''

    return "Unsupported file type."
